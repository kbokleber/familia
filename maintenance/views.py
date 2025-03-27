from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from .models import MaintenanceOrder, Equipment
from django.views.generic import DetailView, ListView, TemplateView
from django.db.models import Q
from datetime import datetime
from django.utils import timezone
from datetime import timedelta
import json
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class CreateOrderView(LoginRequiredMixin, CreateView):
    model = MaintenanceOrder
    template_name = 'maintenance/order_form.html'
    fields = [
        'equipment',
        'status',
        'service_provider',
        'completion_date',
        'cost',
        'description',
        'warranty_expiration',
        'warranty_terms',
        'invoice_number',
        'invoice_file',
        'notes'
    ]
    success_url = reverse_lazy('dashboard:home')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

class CreateEquipmentView(LoginRequiredMixin, CreateView):
    model = Equipment
    template_name = 'maintenance/equipment_form.html'
    fields = ['name', 'type', 'brand', 'model', 'serial_number', 'purchase_date', 'notes']
    success_url = reverse_lazy('dashboard:home')

    def form_valid(self, form):
        form.instance.owner = self.request.user
        return super().form_valid(form)

class DetailOrderView(LoginRequiredMixin, DetailView):
    model = MaintenanceOrder
    template_name = 'maintenance/order_detail.html'
    context_object_name = 'order'

    def get_queryset(self):
        return MaintenanceOrder.objects.all()

class DeleteOrderView(LoginRequiredMixin, DeleteView):
    model = MaintenanceOrder
    success_url = reverse_lazy('dashboard:home')
    template_name = 'maintenance/order_confirm_delete.html'

    def get_queryset(self):
        return MaintenanceOrder.objects.all()

class UpdateOrderView(LoginRequiredMixin, UpdateView):
    model = MaintenanceOrder
    template_name = 'maintenance/order_form.html'
    fields = [
        'equipment',
        'status',
        'service_provider',
        'completion_date',
        'cost',
        'description',
        'warranty_expiration',
        'warranty_terms',
        'invoice_number',
        'invoice_file',
        'notes'
    ]

    def get_success_url(self):
        return reverse_lazy('maintenance:order_detail', kwargs={'pk': self.object.pk})

    def get_queryset(self):
        return MaintenanceOrder.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_update'] = True
        return context

class MaintenanceOrderListView(LoginRequiredMixin, ListView):
    model = MaintenanceOrder
    template_name = 'maintenance/order_list.html'
    context_object_name = 'maintenance_orders'
    paginate_by = 10

    def get_queryset(self):
        queryset = MaintenanceOrder.objects.all()
        
        # Filtro por equipamento
        equipment_id = self.request.GET.get('equipment')
        if equipment_id:
            queryset = queryset.filter(equipment_id=equipment_id)
        
        # Filtro por data
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                queryset = queryset.filter(completion_date__gte=start_date)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(completion_date__lte=end_date)
            except ValueError:
                pass
        
        return queryset.order_by('-completion_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['equipments'] = Equipment.objects.all()
        context['selected_equipment'] = self.request.GET.get('equipment')
        context['start_date'] = self.request.GET.get('start_date')
        context['end_date'] = self.request.GET.get('end_date')
        return context

    def export_to_excel(self, request):
        # Obtém os dados filtrados
        queryset = self.get_queryset()
        
        # Cria um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico de Manutenções"

        # Define os cabeçalhos
        headers = [
            'Equipamento', 'Empresa', 'Data da Manutenção', 'Custo',
            'Garantia até', 'Status', 'Descrição', 'Número da Nota',
            'Observações'
        ]
        
        # Estilo para o cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Adiciona os cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Adiciona os dados
        for row, order in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=str(order.equipment))
            ws.cell(row=row, column=2, value=order.service_provider)
            ws.cell(row=row, column=3, value=order.completion_date.strftime('%d/%m/%Y') if order.completion_date else 'Não definida')
            ws.cell(row=row, column=4, value=f"R$ {order.cost:,.2f}" if order.cost else "R$ 0,00")
            ws.cell(row=row, column=5, value=order.warranty_expiration.strftime('%d/%m/%Y') if order.warranty_expiration else 'Sem garantia')
            ws.cell(row=row, column=6, value=order.get_status_display())
            ws.cell(row=row, column=7, value=order.description or '')
            ws.cell(row=row, column=8, value=order.invoice_number or '')
            ws.cell(row=row, column=9, value=order.notes or '')

        # Ajusta a largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Cria a resposta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="historico_manutencoes.xlsx"'
        
        # Salva o workbook
        wb.save(response)
        
        return response

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'excel':
            return self.export_to_excel(request)
        return super().get(request, *args, **kwargs)

class MaintenanceDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'maintenance/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Contagem de manutenções por status
        context['pending_count'] = MaintenanceOrder.objects.filter(
            status='pendente'
        ).count()

        context['in_progress_count'] = MaintenanceOrder.objects.filter(
            status='em_andamento'
        ).count()

        context['completed_count'] = MaintenanceOrder.objects.filter(
            status='concluida'
        ).count()

        # Total de equipamentos
        context['equipment_count'] = Equipment.objects.count()

        # Últimas 5 manutenções realizadas
        context['recent_maintenance'] = MaintenanceOrder.objects.all().order_by('-completion_date', '-created_at')[:5]

        # Dados para o gráfico de custos
        months = []
        costs = []
        now = timezone.now()
        for i in range(6):
            date = now - timedelta(days=30 * i)
            month_orders = MaintenanceOrder.objects.filter(
                status='concluida',
                completion_date__year=date.year,
                completion_date__month=date.month
            )
            total_cost = sum(order.cost or 0 for order in month_orders)
            months.insert(0, date.strftime('%b/%Y'))
            costs.insert(0, float(total_cost))

        context['months'] = json.dumps(months)
        context['costs'] = json.dumps(costs)
        return context

class EquipmentListView(LoginRequiredMixin, ListView):
    model = Equipment
    template_name = 'maintenance/equipment_list.html'
    context_object_name = 'equipments'
    paginate_by = 10

    def get_queryset(self):
        queryset = Equipment.objects.all()
        
        # Filtro por tipo
        equipment_type = self.request.GET.get('type')
        if equipment_type:
            queryset = queryset.filter(type=equipment_type)
        
        # Filtro por busca
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(brand__icontains=search_query) |
                Q(model__icontains=search_query) |
                Q(serial_number__icontains=search_query)
            )
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['selected_type'] = self.request.GET.get('type')
        context['search_query'] = self.request.GET.get('search')
        context['equipment_types'] = Equipment.TYPES
        return context

class EquipmentDetailView(LoginRequiredMixin, DetailView):
    model = Equipment
    template_name = 'maintenance/equipment_detail.html'
    context_object_name = 'equipment'

    def get_queryset(self):
        return Equipment.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['maintenance_orders'] = MaintenanceOrder.objects.filter(
            equipment=self.object
        ).order_by('-completion_date')
        return context

class EquipmentUpdateView(LoginRequiredMixin, UpdateView):
    model = Equipment
    template_name = 'maintenance/equipment_form.html'
    fields = ['name', 'type', 'brand', 'model', 'serial_number', 'purchase_date', 'notes']
    success_url = reverse_lazy('maintenance:equipment_list')

    def get_queryset(self):
        return Equipment.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_update'] = True
        return context

class EquipmentDeleteView(LoginRequiredMixin, DeleteView):
    model = Equipment
    template_name = 'maintenance/equipment_confirm_delete.html'
    success_url = reverse_lazy('maintenance:equipment_list')

    def get_queryset(self):
        return Equipment.objects.all()